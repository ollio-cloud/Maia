#!/usr/bin/env python3
"""
Complete DNS Audit Tool - Basic + Advanced Checks
Combines SPF/DKIM/DMARC with security and deliverability checks
"""

import dns.resolver
import dns.exception
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import re
import sys
from typing import Dict, List, Tuple
import requests

# Color fills for Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
BOLD_FONT = Font(bold=True)

class CompleteDNSAuditor:
    def __init__(self):
        self.resolver = dns.resolver.Resolver()
        self.resolver.timeout = 5
        self.resolver.lifetime = 5

    # Basic checks (from original audit)
    def check_spf(self, domain: str) -> Tuple[str, str, str]:
        """Check SPF record existence and validity"""
        try:
            answers = self.resolver.resolve(domain, 'TXT')
            spf_records = [str(rdata).strip('"') for rdata in answers if 'v=spf1' in str(rdata)]

            if not spf_records:
                return "FAIL", "No SPF record found", ""

            if len(spf_records) > 1:
                return "WARN", f"Multiple SPF records ({len(spf_records)})", spf_records[0][:100]

            spf = spf_records[0]

            # Check DNS lookup count
            includes = len(re.findall(r'include:', spf))
            a_lookups = len(re.findall(r'\sa\s|\sa:', spf))
            mx_lookups = len(re.findall(r'\smx\s|\smx:', spf))
            total_lookups = includes + a_lookups + mx_lookups

            if total_lookups > 10:
                return "WARN", f"Too many DNS lookups ({total_lookups}/10)", spf[:200]

            if '-all' in spf or '~all' in spf or '?all' in spf:
                return "PASS", f"SPF configured ({len(spf)} chars, {total_lookups} lookups)", spf[:200]
            else:
                return "WARN", "SPF missing fail policy (-all/~all)", spf[:200]

        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.exception.Timeout):
            return "FAIL", "No SPF record", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def check_dkim(self, domain: str) -> Tuple[str, str, str]:
        """Check DKIM records across common selectors"""
        common_selectors = ['selector1', 'selector2', 'google', 's1', 's2', 'default', 'dkim', 'mail']

        found_selectors = []
        for selector in common_selectors:
            dkim_domain = f"{selector}._domainkey.{domain}"
            try:
                answers = self.resolver.resolve(dkim_domain, 'TXT')
                for rdata in answers:
                    record = str(rdata).strip('"')
                    if 'v=DKIM1' in record or 'p=' in record:
                        key_match = re.search(r'p=([A-Za-z0-9+/]+)', record)
                        if key_match:
                            key_b64 = key_match.group(1)
                            key_bits = len(key_b64) * 6
                            found_selectors.append(f"{selector}:{key_bits}b")
                        else:
                            found_selectors.append(selector)
                        break
            except:
                continue

        if not found_selectors:
            return "FAIL", "No DKIM records found", ""

        if len(found_selectors) >= 2:
            return "PASS", f"DKIM configured ({len(found_selectors)} selectors)", ", ".join(found_selectors[:3])
        else:
            return "WARN", f"DKIM found (1 selector only)", ", ".join(found_selectors)

    def check_dmarc(self, domain: str) -> Tuple[str, str, str]:
        """Check DMARC record"""
        dmarc_domain = f"_dmarc.{domain}"
        try:
            answers = self.resolver.resolve(dmarc_domain, 'TXT')
            dmarc_records = [str(rdata).strip('"') for rdata in answers if 'v=DMARC1' in str(rdata)]

            if not dmarc_records:
                return "FAIL", "No DMARC record", ""

            dmarc = dmarc_records[0]

            if 'p=reject' in dmarc:
                return "PASS", "DMARC p=reject (best)", dmarc[:200]
            elif 'p=quarantine' in dmarc:
                return "PASS", "DMARC p=quarantine (good)", dmarc[:200]
            elif 'p=none' in dmarc:
                return "WARN", "DMARC p=none (monitor only)", dmarc[:200]
            else:
                return "WARN", "DMARC policy unclear", dmarc[:200]

        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.exception.Timeout):
            return "FAIL", "No DMARC record", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def check_mx(self, domain: str) -> Tuple[str, str, str]:
        """Check MX records"""
        try:
            answers = self.resolver.resolve(domain, 'MX')
            mx_records = sorted([(rdata.preference, str(rdata.exchange)) for rdata in answers])

            if not mx_records:
                return "FAIL", "No MX records", ""

            mx_list = [f"{pref}:{host}" for pref, host in mx_records[:3]]
            return "PASS", f"{len(mx_records)} MX record(s)", ", ".join(mx_list)[:100]

        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer):
            return "WARN", "No MX records (no email)", ""
        except dns.exception.Timeout:
            return "ERROR", "DNS timeout", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def check_ns(self, domain: str) -> Tuple[str, str, str]:
        """Check NS records"""
        try:
            answers = self.resolver.resolve(domain, 'NS')
            ns_records = [str(rdata) for rdata in answers]

            if not ns_records:
                return "FAIL", "No NS records", ""

            route53_count = sum(1 for ns in ns_records if 'awsdns' in ns.lower())

            if route53_count > 0:
                return "PASS", f"{len(ns_records)} NS (Route53)", ", ".join(ns_records[:2])[:100]
            else:
                return "PASS", f"{len(ns_records)} NS", ", ".join(ns_records[:2])[:100]

        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.exception.Timeout):
            return "FAIL", "No NS records", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    # Advanced security checks
    def check_caa(self, domain: str) -> Tuple[str, str]:
        """Check CAA records"""
        try:
            answers = self.resolver.resolve(domain, 'CAA')
            caa_records = [f"{rdata.flags} {rdata.tag} {rdata.value}" for rdata in answers]
            if caa_records:
                return "PASS", f"CAA configured ({len(caa_records)})"
            return "WARN", "No CAA records"
        except:
            return "WARN", "No CAA records"

    def check_mta_sts(self, domain: str) -> Tuple[str, str]:
        """Check MTA-STS"""
        try:
            answers = self.resolver.resolve(f"_mta-sts.{domain}", 'TXT')
            for rdata in answers:
                if 'v=STSv1' in str(rdata):
                    return "PASS", "MTA-STS configured"
            return "WARN", "No MTA-STS"
        except:
            return "WARN", "No MTA-STS"

    def check_subdomain_takeover(self, domain: str) -> Tuple[str, str]:
        """Quick subdomain takeover check"""
        vulnerable_patterns = ['herokuapp.com', 'github.io', 'azurewebsites.net', 's3.amazonaws.com']
        try:
            for subdomain in ['www', 'blog', 'dev']:
                try:
                    answers = self.resolver.resolve(f"{subdomain}.{domain}", 'CNAME')
                    for rdata in answers:
                        target = str(rdata.target).lower()
                        for pattern in vulnerable_patterns:
                            if pattern in target:
                                return "FAIL", f"Potential takeover: {subdomain}"
                except:
                    continue
            return "PASS", "No obvious risks"
        except:
            return "INFO", "Check inconclusive"

    def audit_domain_complete(self, domain: str) -> Dict:
        """Run complete DNS audit"""
        domain = domain.strip().rstrip('.')
        print(f"Auditing: {domain}")

        # Basic checks
        ns_status, ns_note, ns_detail = self.check_ns(domain)
        mx_status, mx_note, mx_detail = self.check_mx(domain)
        spf_status, spf_note, spf_detail = self.check_spf(domain)
        dkim_status, dkim_note, dkim_detail = self.check_dkim(domain)
        dmarc_status, dmarc_note, dmarc_detail = self.check_dmarc(domain)

        # Advanced checks
        caa_status, caa_note = self.check_caa(domain)
        mta_sts_status, mta_sts_note = self.check_mta_sts(domain)
        subdomain_status, subdomain_note = self.check_subdomain_takeover(domain)

        # Calculate overall status
        critical_statuses = [ns_status, spf_status, dkim_status, dmarc_status]
        security_statuses = [caa_status, subdomain_status]

        if "FAIL" in critical_statuses or ns_status == "FAIL":
            overall = "FAIL"
        elif subdomain_status == "FAIL":
            overall = "HIGH RISK"
        elif "ERROR" in critical_statuses:
            overall = "ERROR"
        elif "WARN" in critical_statuses:
            overall = "WARN"
        else:
            overall = "PASS"

        # Security score
        security_score = 0
        if caa_status == "PASS": security_score += 25
        if mta_sts_status == "PASS": security_score += 25
        if subdomain_status == "PASS": security_score += 25
        if dmarc_status == "PASS" and 'reject' in dmarc_note.lower(): security_score += 25

        return {
            'domain': domain,
            'overall_status': overall,
            'security_score': security_score,
            'ns_status': ns_status,
            'ns_note': ns_note,
            'ns_detail': ns_detail,
            'mx_status': mx_status,
            'mx_note': mx_note,
            'mx_detail': mx_detail,
            'spf_status': spf_status,
            'spf_note': spf_note,
            'spf_detail': spf_detail,
            'dkim_status': dkim_status,
            'dkim_note': dkim_note,
            'dkim_detail': dkim_detail,
            'dmarc_status': dmarc_status,
            'dmarc_note': dmarc_note,
            'dmarc_detail': dmarc_detail,
            'caa_status': caa_status,
            'caa_note': caa_note,
            'mta_sts_status': mta_sts_status,
            'mta_sts_note': mta_sts_note,
            'subdomain_status': subdomain_status,
            'subdomain_note': subdomain_note,
        }

def update_excel_complete(input_file: str, output_file: str, results: List[Dict]):
    """Update Excel with complete audit results"""
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    headers = ['Domain', 'Overall Status', 'Security Score',
               'NS Status', 'NS Note', 'NS Details',
               'MX Status', 'MX Note', 'MX Details',
               'SPF Status', 'SPF Note', 'SPF Details',
               'DKIM Status', 'DKIM Note', 'DKIM Details',
               'DMARC Status', 'DMARC Note', 'DMARC Details',
               'CAA Status', 'CAA Note',
               'MTA-STS Status', 'MTA-STS Note',
               'Subdomain Risk', 'Subdomain Note',
               'Audit Date']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD_FONT

    for row_idx, result in enumerate(results, start=2):
        col = 1
        ws.cell(row=row_idx, column=col, value=result['domain']); col += 1

        # Overall status
        overall_cell = ws.cell(row=row_idx, column=col, value=result['overall_status'])
        if result['overall_status'] == 'PASS':
            overall_cell.fill = GREEN_FILL
        elif result['overall_status'] in ['FAIL', 'HIGH RISK']:
            overall_cell.fill = RED_FILL
        elif result['overall_status'] == 'WARN':
            overall_cell.fill = YELLOW_FILL
        col += 1

        # Security score
        score_cell = ws.cell(row=row_idx, column=col, value=result['security_score'])
        if result['security_score'] >= 75:
            score_cell.fill = GREEN_FILL
        elif result['security_score'] >= 50:
            score_cell.fill = YELLOW_FILL
        else:
            score_cell.fill = RED_FILL
        col += 1

        # NS, MX, SPF, DKIM, DMARC (with details)
        for check in ['ns', 'mx', 'spf', 'dkim', 'dmarc']:
            status_cell = ws.cell(row=row_idx, column=col, value=result[f'{check}_status'])
            if result[f'{check}_status'] == 'PASS':
                status_cell.fill = GREEN_FILL
            elif result[f'{check}_status'] == 'FAIL':
                status_cell.fill = RED_FILL
            elif result[f'{check}_status'] == 'WARN':
                status_cell.fill = YELLOW_FILL
            col += 1
            ws.cell(row=row_idx, column=col, value=result[f'{check}_note']); col += 1
            ws.cell(row=row_idx, column=col, value=result[f'{check}_detail']); col += 1

        # Advanced checks (CAA, MTA-STS, Subdomain)
        for check in ['caa', 'mta_sts', 'subdomain']:
            status_cell = ws.cell(row=row_idx, column=col, value=result[f'{check}_status'])
            if result[f'{check}_status'] == 'PASS':
                status_cell.fill = GREEN_FILL
            elif result[f'{check}_status'] == 'FAIL':
                status_cell.fill = RED_FILL
            elif result[f'{check}_status'] == 'WARN':
                status_cell.fill = YELLOW_FILL
            col += 1
            ws.cell(row=row_idx, column=col, value=result[f'{check}_note']); col += 1

        # Audit date
        ws.cell(row=row_idx, column=col, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

    # Adjust column widths
    for i, width in enumerate([40, 15, 12, 12, 30, 50, 12, 30, 50, 12, 30, 80, 12, 30, 50, 12, 30, 80, 12, 25, 12, 25, 12, 30, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    wb.save(output_file)
    print(f"\n✅ Results saved to: {output_file}")

def main():
    input_file = '/Users/YOUR_USERNAME/Library/CloudStorage/OneDrive-YOUR_ORG/Documents/Claude/Route53 domains - 2025.10.10 1.xlsx'
    output_file = '/Users/YOUR_USERNAME/Library/CloudStorage/OneDrive-YOUR_ORG/Documents/Claude/Route53 domains - Complete Audit.xlsx'

    # Load domains
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    domains = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            domain = str(row[0]).strip().rstrip('.')
            domains.append(domain)

    print(f"📋 Found {len(domains)} domains for complete audit\n")

    # Run complete audit
    auditor = CompleteDNSAuditor()
    results = []

    for idx, domain in enumerate(domains, start=1):
        print(f"[{idx}/{len(domains)}] ", end="")
        result = auditor.audit_domain_complete(domain)
        results.append(result)

    # Update Excel
    print(f"\n📝 Updating Excel with complete audit results...")
    update_excel_complete(input_file, output_file, results)

    # Summary
    overall_pass = sum(1 for r in results if r['overall_status'] == 'PASS')
    overall_warn = sum(1 for r in results if r['overall_status'] == 'WARN')
    overall_fail = sum(1 for r in results if r['overall_status'] == 'FAIL')
    high_risk = sum(1 for r in results if r['overall_status'] == 'HIGH RISK')

    avg_security_score = sum(r['security_score'] for r in results) / len(results)

    print(f"\n📊 Complete DNS Audit Summary:")
    print(f"   ✅ PASS:      {overall_pass}/{len(domains)} ({overall_pass*100//len(domains)}%)")
    print(f"   ⚠️  WARN:      {overall_warn}/{len(domains)} ({overall_warn*100//len(domains)}%)")
    print(f"   ❌ FAIL:      {overall_fail}/{len(domains)} ({overall_fail*100//len(domains)}%)")
    print(f"   🔴 HIGH RISK: {high_risk}/{len(domains)} ({high_risk*100//len(domains) if domains else 0}%)")
    print(f"\n🔒 Average Security Score: {avg_security_score:.1f}/100")

    # Advanced findings
    caa_missing = sum(1 for r in results if r['caa_status'] != 'PASS')
    mta_sts_missing = sum(1 for r in results if r['mta_sts_status'] != 'PASS')
    subdomain_risk = sum(1 for r in results if r['subdomain_status'] == 'FAIL')

    print(f"\n🔍 Advanced Security Issues:")
    print(f"   CAA Records missing:      {caa_missing} domains ({caa_missing*100//len(domains)}%)")
    print(f"   MTA-STS not configured:   {mta_sts_missing} domains ({mta_sts_missing*100//len(domains)}%)")
    print(f"   Subdomain takeover risk:  {subdomain_risk} domains ({subdomain_risk*100//len(domains) if domains else 0}%)")

if __name__ == '__main__':
    main()
