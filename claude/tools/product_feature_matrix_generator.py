#!/usr/bin/env python3
"""
Product Feature Matrix Generator
Creates detailed Excel spreadsheet with feature comparison across products

Created: 2025-10-06
Purpose: Visual feature matrix for portfolio analysis
"""

import json
from pathlib import Path
from typing import Dict, List, Set
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ProductFeatureMatrixGenerator:
    """Generate Excel feature comparison matrix"""

    def __init__(self, data_dir: str = None):
        if data_dir is None:
            maia_root = Path(__file__).parent.parent.parent
            self.data_dir = maia_root / "claude" / "data" / "product_intelligence"
        else:
            self.data_dir = Path(data_dir)

        self.products = {}
        self.load_products()

    def load_products(self):
        """Load all product JSON files"""
        if not self.data_dir.exists():
            print(f"❌ Data directory not found: {self.data_dir}")
            return

        for json_file in self.data_dir.glob("*.json"):
            try:
                with open(json_file, 'r') as f:
                    product_data = json.load(f)
                    product_name = product_data.get('product_name', json_file.stem)
                    self.products[product_name] = product_data
                    print(f"✅ Loaded: {product_name}")
            except Exception as e:
                print(f"❌ Error loading {json_file}: {e}")

    def extract_all_features_by_category(self) -> Dict[str, Dict[str, Set[str]]]:
        """Extract all features organized by category and product"""
        # Structure: {category: {feature_name: {product1, product2, ...}}}
        feature_matrix = defaultdict(lambda: defaultdict(set))

        for product_name, product_data in self.products.items():
            capabilities = product_data.get('capabilities', {})

            for cap_name, cap_data in capabilities.items():
                features = cap_data.get('features', [])

                for feature in features:
                    feature_name = feature.get('name', '')
                    if feature_name:
                        feature_matrix[cap_name][feature_name].add(product_name)

        return dict(feature_matrix)

    def create_feature_matrix_excel(self, output_file: str = None):
        """Generate Excel spreadsheet with feature matrix"""
        if output_file is None:
            output_file = self.data_dir / "Product_Feature_Matrix.xlsx"
        else:
            output_file = Path(output_file)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feature Matrix"

        # Styling
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        category_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        category_font = Font(color="FFFFFF", bold=True, size=12)
        feature_font = Font(size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get feature matrix
        feature_matrix = self.extract_all_features_by_category()
        product_names = sorted(self.products.keys())

        # Column widths
        ws.column_dimensions['A'].width = 50  # Feature name column
        ws.column_dimensions['B'].width = 20  # Category column
        for i, _ in enumerate(product_names, start=3):
            ws.column_dimensions[get_column_letter(i)].width = 18

        # Header row
        ws['A1'] = "Feature Name"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align
        ws['A1'].border = thin_border

        ws['B1'] = "Category"
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
        ws['B1'].alignment = center_align
        ws['B1'].border = thin_border

        # Product headers
        for idx, product_name in enumerate(product_names, start=3):
            cell = ws.cell(row=1, column=idx)
            cell.value = product_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Current row tracker
        current_row = 2

        # Group features by category
        sorted_categories = sorted(feature_matrix.keys())

        for category in sorted_categories:
            features = feature_matrix[category]

            # Category header row
            category_cell = ws.cell(row=current_row, column=1)
            category_cell.value = f"📦 {category}"
            category_cell.font = category_font
            category_cell.fill = category_fill
            category_cell.alignment = left_align
            category_cell.border = thin_border

            # Merge category header across all columns
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(product_names) + 2)

            current_row += 1

            # Add features for this category
            sorted_features = sorted(features.keys())

            for feature_name in sorted_features:
                products_with_feature = features[feature_name]

                # Feature name
                feature_cell = ws.cell(row=current_row, column=1)
                feature_cell.value = feature_name
                feature_cell.font = feature_font
                feature_cell.alignment = left_align
                feature_cell.border = thin_border

                # Category name
                cat_cell = ws.cell(row=current_row, column=2)
                cat_cell.value = category
                cat_cell.font = Font(size=9, italic=True)
                cat_cell.alignment = center_align
                cat_cell.border = thin_border

                # Product checkmarks
                for idx, product_name in enumerate(product_names, start=3):
                    cell = ws.cell(row=current_row, column=idx)

                    if product_name in products_with_feature:
                        cell.value = "✓"
                        cell.font = Font(size=14, bold=True, color="008000")
                        cell.alignment = center_align
                    else:
                        cell.value = ""

                    cell.border = thin_border

                current_row += 1

            # Add blank row between categories
            current_row += 1

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.column_dimensions['A'].width = 40
        summary_ws.column_dimensions['B'].width = 15

        summary_ws['A1'] = "Product"
        summary_ws['A1'].font = header_font
        summary_ws['A1'].fill = header_fill
        summary_ws['B1'] = "Total Features"
        summary_ws['B1'].font = header_font
        summary_ws['B1'].fill = header_fill

        row = 2
        for product_name in product_names:
            feature_count = 0
            for category_features in feature_matrix.values():
                for products in category_features.values():
                    if product_name in products:
                        feature_count += 1

            summary_ws.cell(row=row, column=1).value = product_name
            summary_ws.cell(row=row, column=2).value = feature_count
            summary_ws.cell(row=row, column=2).alignment = center_align
            row += 1

        # Add category summary
        summary_ws['A' + str(row + 1)] = "Category"
        summary_ws['A' + str(row + 1)].font = header_font
        summary_ws['A' + str(row + 1)].fill = header_fill
        summary_ws['B' + str(row + 1)] = "Unique Features"
        summary_ws['B' + str(row + 1)].font = header_font
        summary_ws['B' + str(row + 1)].fill = header_fill

        row = row + 2
        for category in sorted_categories:
            summary_ws.cell(row=row, column=1).value = category
            summary_ws.cell(row=row, column=2).value = len(feature_matrix[category])
            summary_ws.cell(row=row, column=2).alignment = center_align
            row += 1

        # Freeze panes
        ws.freeze_panes = 'A2'
        summary_ws.freeze_panes = 'A2'

        # Save workbook
        wb.save(output_file)
        print(f"\n✅ Feature matrix saved to: {output_file}")

        return output_file

    def generate_overlap_summary(self) -> Dict:
        """Generate overlap statistics"""
        feature_matrix = self.extract_all_features_by_category()

        stats = {
            'total_categories': len(feature_matrix),
            'total_unique_features': 0,
            'features_in_multiple_products': 0,
            'features_in_single_product': 0,
            'overlap_details': []
        }

        for category, features in feature_matrix.items():
            for feature_name, products in features.items():
                stats['total_unique_features'] += 1

                if len(products) > 1:
                    stats['features_in_multiple_products'] += 1
                    stats['overlap_details'].append({
                        'feature': feature_name,
                        'category': category,
                        'products': list(products),
                        'count': len(products)
                    })
                else:
                    stats['features_in_single_product'] += 1

        # Sort overlap details by product count (descending)
        stats['overlap_details'].sort(key=lambda x: x['count'], reverse=True)

        return stats


def main():
    """CLI interface for feature matrix generation"""
    import argparse

    parser = argparse.ArgumentParser(description="Generate Excel feature comparison matrix")
    parser.add_argument('--output', type=str, help='Output Excel file path')
    parser.add_argument('--stats', action='store_true', help='Show overlap statistics')

    args = parser.parse_args()

    generator = ProductFeatureMatrixGenerator()

    if not generator.products:
        print("❌ No products loaded. Check data directory.")
        return

    print(f"\n✅ Loaded {len(generator.products)} products\n")

    if args.stats:
        stats = generator.generate_overlap_summary()
        print(f"📊 Statistics:")
        print(f"  Total Categories: {stats['total_categories']}")
        print(f"  Total Unique Features: {stats['total_unique_features']}")
        print(f"  Features in Single Product: {stats['features_in_single_product']}")
        print(f"  Features in Multiple Products: {stats['features_in_multiple_products']}")
        print(f"\n🔍 Top Overlapping Features:")
        for detail in stats['overlap_details'][:10]:
            print(f"  • {detail['feature']} ({detail['category']})")
            print(f"    Products: {', '.join(detail['products'])} ({detail['count']} products)")
    else:
        generator.create_feature_matrix_excel(output_file=args.output)

if __name__ == "__main__":
    main()
