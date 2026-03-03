#!/usr/bin/env python3
"""
Intelligent Excel Feature Matrix Generator
Creates enhanced feature comparison with semantic similarity grouping

Created: 2025-10-06
Purpose: Visual feature matrix showing true overlaps via intelligent matching
"""

import json
from pathlib import Path
from typing import Dict, List
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class IntelligentExcelMatrixGenerator:
    """Generate Excel with intelligent feature grouping"""

    def __init__(self, data_dir: str = None):
        if data_dir is None:
            maia_root = Path(__file__).parent.parent.parent
            self.data_dir = maia_root / "claude" / "data" / "product_intelligence"
        else:
            self.data_dir = Path(data_dir)

        self.products = {}
        self.similarity_data = None
        self.load_data()

    def load_data(self):
        """Load product data and similarity analysis"""
        # Load products
        for json_file in self.data_dir.glob("*.json"):
            if json_file.name == "feature_similarity_analysis.json":
                continue
            try:
                with open(json_file, 'r') as f:
                    product_data = json.load(f)
                    product_name = product_data.get('product_name', json_file.stem)
                    self.products[product_name] = product_data
            except Exception as e:
                print(f"❌ Error loading {json_file}: {e}")

        # Load similarity analysis
        similarity_file = self.data_dir / "feature_similarity_analysis.json"
        if similarity_file.exists():
            with open(similarity_file, 'r') as f:
                self.similarity_data = json.load(f)
                print(f"✅ Loaded similarity analysis: {len(self.similarity_data['overlaps'])} overlaps")

    def group_similar_features(self) -> List[Dict]:
        """Group features by similarity clusters"""
        if not self.similarity_data:
            print("⚠️  No similarity data available")
            return []

        # Create feature groups
        feature_groups = []
        processed_features = set()

        for overlap in self.similarity_data['overlaps']:
            if overlap['similarity'] >= 0.75:  # High and medium similarity
                f1_key = f"{overlap['feature1']['product']}:{overlap['feature1']['name']}"
                f2_key = f"{overlap['feature2']['product']}:{overlap['feature2']['name']}"

                if f1_key not in processed_features or f2_key not in processed_features:
                    group = {
                        'similarity': overlap['similarity'],
                        'features': [overlap['feature1'], overlap['feature2']],
                        'group_name': overlap['feature1']['name']  # Use first feature as representative
                    }
                    feature_groups.append(group)
                    processed_features.add(f1_key)
                    processed_features.add(f2_key)

        return feature_groups

    def create_enhanced_matrix(self, output_file: str = None):
        """Generate enhanced Excel with intelligent grouping"""
        if output_file is None:
            output_file = self.data_dir / "Product_Feature_Matrix_Intelligent.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Intelligent Feature Matrix"

        # Styling
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        overlap_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for overlaps
        unique_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for unique
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        product_names = sorted(self.products.keys())

        # Column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        for i, _ in enumerate(product_names, start=3):
            ws.column_dimensions[get_column_letter(i)].width = 15

        # Headers
        ws['A1'] = "Feature Name"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = center_align

        ws['B1'] = "Overlap Count"
        ws['B1'].font = header_font
        ws['B1'].fill = header_fill
        ws['B1'].alignment = center_align

        for idx, product_name in enumerate(product_names, start=3):
            cell = ws.cell(row=1, column=idx)
            cell.value = product_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        current_row = 2

        # Section 1: High Overlap Features (≥75%)
        if self.similarity_data:
            high_overlap_groups = [g for g in self.group_similar_features() if g['similarity'] >= 0.75]

            if high_overlap_groups:
                # Section header
                section_cell = ws.cell(row=current_row, column=1)
                section_cell.value = "🔴 HIGH OVERLAP FEATURES (≥75% Similarity) - Consolidation Candidates"
                section_cell.font = Font(bold=True, size=12, color="FFFFFF")
                section_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
                ws.merge_cells(start_row=current_row, start_column=1,
                              end_row=current_row, end_column=len(product_names) + 2)
                current_row += 1

                for group in sorted(high_overlap_groups, key=lambda x: x['similarity'], reverse=True):
                    # Feature name (using first feature as representative)
                    feature_cell = ws.cell(row=current_row, column=1)
                    feature_names = [f['name'] for f in group['features']]
                    feature_cell.value = f"{group['features'][0]['name']} ({group['similarity']:.0%} match)"
                    feature_cell.font = Font(bold=True)
                    feature_cell.fill = overlap_fill
                    feature_cell.alignment = left_align

                    # Overlap count
                    overlap_cell = ws.cell(row=current_row, column=2)
                    overlap_cell.value = len(group['features'])
                    overlap_cell.fill = overlap_fill
                    overlap_cell.alignment = center_align

                    # Product checkmarks
                    products_in_group = {f['product'] for f in group['features']}
                    for idx, product_name in enumerate(product_names, start=3):
                        cell = ws.cell(row=current_row, column=idx)
                        if product_name in products_in_group:
                            cell.value = "✓"
                            cell.font = Font(size=14, bold=True, color="C00000")
                            cell.fill = overlap_fill
                        cell.alignment = center_align

                    current_row += 1

                current_row += 1  # Blank row

        # Section 2: All Features by Category (including unique)
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.value = "✅ ALL FEATURES BY CATEGORY"
        section_cell.font = Font(bold=True, size=12, color="FFFFFF")
        section_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells(start_row=current_row, start_column=1,
                      end_row=current_row, end_column=len(product_names) + 2)
        current_row += 1

        # Extract all features by category
        feature_matrix = defaultdict(lambda: defaultdict(set))
        for product_name, product_data in self.products.items():
            capabilities = product_data.get('capabilities', {})
            for cap_name, cap_data in capabilities.items():
                features = cap_data.get('features', [])
                for feature in features:
                    feature_name = feature.get('name', '')
                    if feature_name:
                        feature_matrix[cap_name][feature_name].add(product_name)

        for category in sorted(feature_matrix.keys()):
            features = feature_matrix[category]

            # Category header
            category_cell = ws.cell(row=current_row, column=1)
            category_cell.value = f"📦 {category}"
            category_cell.font = Font(bold=True, size=11, color="FFFFFF")
            category_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=len(product_names) + 2)
            current_row += 1

            for feature_name in sorted(features.keys()):
                products_with_feature = features[feature_name]

                # Feature name
                feature_cell = ws.cell(row=current_row, column=1)
                feature_cell.value = feature_name
                feature_cell.alignment = left_align

                # Overlap count
                overlap_count = len(products_with_feature)
                overlap_cell = ws.cell(row=current_row, column=2)
                overlap_cell.value = overlap_count if overlap_count > 1 else ""
                overlap_cell.alignment = center_align

                # Color code based on overlap
                if overlap_count > 1:
                    feature_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                    overlap_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    feature_cell.fill = unique_fill
                    overlap_cell.fill = unique_fill

                # Product checkmarks
                for idx, product_name in enumerate(product_names, start=3):
                    cell = ws.cell(row=current_row, column=idx)
                    if product_name in products_with_feature:
                        cell.value = "✓"
                        if overlap_count > 1:
                            cell.font = Font(size=14, bold=True, color="FF6600")
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        else:
                            cell.font = Font(size=14, bold=True, color="008000")
                            cell.fill = unique_fill
                    cell.alignment = center_align

                current_row += 1

            current_row += 1  # Blank row between categories

        # Freeze panes
        ws.freeze_panes = 'A2'

        # Save
        wb.save(output_file)
        print(f"✅ Intelligent feature matrix saved to: {output_file}")
        return output_file


def main():
    """CLI interface"""
    import argparse

    parser = argparse.ArgumentParser(description="Generate intelligent Excel feature matrix")
    parser.add_argument('--output', type=str, help='Output Excel file path')

    args = parser.parse_args()

    generator = IntelligentExcelMatrixGenerator()

    if not generator.products:
        print("❌ No products loaded")
        return

    generator.create_enhanced_matrix(output_file=args.output)


if __name__ == "__main__":
    main()
