#!/usr/bin/env python3
"""
DNS Audit Tool for Route53 Domains
Checks SPF, DKIM, DMARC, MX, NS records and reports pass/fail status
"""

import dns.resolver
import dns.exception
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import re
import sys
from typing import Dict, List, Tuple

# Color fills for Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BOLD_FONT = Font(bold=True)

class DNSAuditor:
    def __init__(self):
        self.resolver = dns.resolver.Resolver()
        self.resolver.timeout = 5
        self.resolver.lifetime = 5

    def check_spf(self, domain: str) -> Tuple[str, str, str]:
        """Check SPF record existence and validity"""
        try:
            answers = self.resolver.resolve(domain, 'TXT')
            spf_records = [str(rdata).strip('"') for rdata in answers if 'v=spf1' in str(rdata)]

            if not spf_records:
                return "FAIL", "No SPF record found", ""

            if len(spf_records) > 1:
                return "WARN", f"Multiple SPF records ({len(spf_records)})", spf_records[0][:100]

            spf = spf_records[0]

            # Check DNS lookup count (rough estimate)
            includes = len(re.findall(r'include:', spf))
            a_lookups = len(re.findall(r'\sa\s|\sa:', spf))
            mx_lookups = len(re.findall(r'\smx\s|\smx:', spf))
            total_lookups = includes + a_lookups + mx_lookups

            if total_lookups > 10:
                return "WARN", f"Too many DNS lookups ({total_lookups}/10)", spf[:100]

            if '-all' in spf or '~all' in spf or '?all' in spf:
                return "PASS", f"SPF configured ({len(spf)} chars)", spf[:100]
            else:
                return "WARN", "SPF missing fail policy (-all/~all)", spf[:100]

        except dns.resolver.NXDOMAIN:
            return "FAIL", "Domain does not exist", ""
        except dns.resolver.NoAnswer:
            return "FAIL", "No TXT records", ""
        except dns.exception.Timeout:
            return "ERROR", "DNS timeout", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def check_dkim(self, domain: str) -> Tuple[str, str, str]:
        """Check DKIM records across common selectors"""
        # Common DKIM selectors to probe
        common_selectors = [
            'selector1', 'selector2',  # Microsoft 365
            'google',                   # Google Workspace
            's1', 's2',                # SendGrid
            'default',                 # Generic
            'k1', 'k2',                # Some mail servers
            'dkim',                    # Generic
            'mail',                    # Generic
            'email',                   # Generic
        ]

        found_selectors = []

        for selector in common_selectors:
            dkim_domain = f"{selector}._domainkey.{domain}"
            try:
                answers = self.resolver.resolve(dkim_domain, 'TXT')
                for rdata in answers:
                    record = str(rdata).strip('"')
                    if 'v=DKIM1' in record or 'p=' in record:
                        # Check key length (rough estimate from public key)
                        key_match = re.search(r'p=([A-Za-z0-9+/]+)', record)
                        if key_match:
                            key_b64 = key_match.group(1)
                            key_bits = len(key_b64) * 6  # Rough estimate
                            found_selectors.append(f"{selector}:{key_bits}b")
                        else:
                            found_selectors.append(selector)
                        break
            except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.exception.Timeout):
                continue
            except Exception:
                continue

        if not found_selectors:
            return "FAIL", "No DKIM records found", ""

        if len(found_selectors) >= 2:
            return "PASS", f"DKIM configured ({len(found_selectors)} selectors)", ", ".join(found_selectors[:3])
        else:
            return "WARN", f"DKIM found (1 selector only)", ", ".join(found_selectors)

    def check_dmarc(self, domain: str) -> Tuple[str, str, str]:
        """Check DMARC record"""
        dmarc_domain = f"_dmarc.{domain}"
        try:
            answers = self.resolver.resolve(dmarc_domain, 'TXT')
            dmarc_records = [str(rdata).strip('"') for rdata in answers if 'v=DMARC1' in str(rdata)]

            if not dmarc_records:
                return "FAIL", "No DMARC record", ""

            dmarc = dmarc_records[0]

            # Check policy
            if 'p=reject' in dmarc:
                return "PASS", "DMARC p=reject (best)", dmarc[:100]
            elif 'p=quarantine' in dmarc:
                return "PASS", "DMARC p=quarantine (good)", dmarc[:100]
            elif 'p=none' in dmarc:
                return "WARN", "DMARC p=none (monitor only)", dmarc[:100]
            else:
                return "WARN", "DMARC policy unclear", dmarc[:100]

        except dns.resolver.NXDOMAIN:
            return "FAIL", "No DMARC record", ""
        except dns.resolver.NoAnswer:
            return "FAIL", "No DMARC record", ""
        except dns.exception.Timeout:
            return "ERROR", "DNS timeout", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def check_mx(self, domain: str) -> Tuple[str, str, str]:
        """Check MX records"""
        try:
            answers = self.resolver.resolve(domain, 'MX')
            mx_records = sorted([(rdata.preference, str(rdata.exchange)) for rdata in answers])

            if not mx_records:
                return "FAIL", "No MX records", ""

            mx_list = [f"{pref}:{host}" for pref, host in mx_records[:3]]
            return "PASS", f"{len(mx_records)} MX record(s)", ", ".join(mx_list)[:100]

        except dns.resolver.NXDOMAIN:
            return "FAIL", "Domain does not exist", ""
        except dns.resolver.NoAnswer:
            return "WARN", "No MX records (no email)", ""
        except dns.exception.Timeout:
            return "ERROR", "DNS timeout", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def check_ns(self, domain: str) -> Tuple[str, str, str]:
        """Check NS records"""
        try:
            answers = self.resolver.resolve(domain, 'NS')
            ns_records = [str(rdata) for rdata in answers]

            if not ns_records:
                return "FAIL", "No NS records", ""

            # Check if using Route53
            route53_count = sum(1 for ns in ns_records if 'awsdns' in ns.lower())

            if route53_count > 0:
                return "PASS", f"{len(ns_records)} NS (Route53)", ", ".join(ns_records[:2])[:100]
            else:
                return "PASS", f"{len(ns_records)} NS", ", ".join(ns_records[:2])[:100]

        except dns.resolver.NXDOMAIN:
            return "FAIL", "Domain does not exist", ""
        except dns.resolver.NoAnswer:
            return "FAIL", "No NS records", ""
        except dns.exception.Timeout:
            return "ERROR", "DNS timeout", ""
        except Exception as e:
            return "ERROR", f"Error: {str(e)[:50]}", ""

    def audit_domain(self, domain: str) -> Dict:
        """Run complete DNS audit for domain"""
        domain = domain.strip().rstrip('.')

        print(f"Auditing: {domain}")

        # Check all DNS records
        ns_status, ns_note, ns_detail = self.check_ns(domain)
        mx_status, mx_note, mx_detail = self.check_mx(domain)
        spf_status, spf_note, spf_detail = self.check_spf(domain)
        dkim_status, dkim_note, dkim_detail = self.check_dkim(domain)
        dmarc_status, dmarc_note, dmarc_detail = self.check_dmarc(domain)

        # Calculate overall status
        statuses = [ns_status, mx_status, spf_status, dkim_status, dmarc_status]
        if "FAIL" in statuses or ns_status == "FAIL":
            overall = "FAIL"
        elif "ERROR" in statuses:
            overall = "ERROR"
        elif "WARN" in statuses:
            overall = "WARN"
        else:
            overall = "PASS"

        return {
            'domain': domain,
            'overall_status': overall,
            'ns_status': ns_status,
            'ns_note': ns_note,
            'ns_detail': ns_detail,
            'mx_status': mx_status,
            'mx_note': mx_note,
            'mx_detail': mx_detail,
            'spf_status': spf_status,
            'spf_note': spf_note,
            'spf_detail': spf_detail,
            'dkim_status': dkim_status,
            'dkim_note': dkim_note,
            'dkim_detail': dkim_detail,
            'dmarc_status': dmarc_status,
            'dmarc_note': dmarc_note,
            'dmarc_detail': dmarc_detail,
        }

def update_excel_with_results(input_file: str, output_file: str, results: List[Dict]):
    """Update Excel file with DNS audit results"""

    # Load workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Add headers if not present
    headers = ['Domain', 'Overall Status',
               'NS Status', 'NS Note', 'NS Details',
               'MX Status', 'MX Note', 'MX Details',
               'SPF Status', 'SPF Note', 'SPF Details',
               'DKIM Status', 'DKIM Note', 'DKIM Details',
               'DMARC Status', 'DMARC Note', 'DMARC Details',
               'Audit Date']

    # Clear existing headers and add new ones
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = BOLD_FONT

    # Write results
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=result['domain'])

        # Overall status
        overall_cell = ws.cell(row=row_idx, column=2, value=result['overall_status'])
        if result['overall_status'] == 'PASS':
            overall_cell.fill = GREEN_FILL
        elif result['overall_status'] == 'FAIL':
            overall_cell.fill = RED_FILL
        elif result['overall_status'] == 'WARN':
            overall_cell.fill = YELLOW_FILL

        # NS
        ns_status_cell = ws.cell(row=row_idx, column=3, value=result['ns_status'])
        ws.cell(row=row_idx, column=4, value=result['ns_note'])
        ws.cell(row=row_idx, column=5, value=result['ns_detail'])
        if result['ns_status'] == 'PASS':
            ns_status_cell.fill = GREEN_FILL
        elif result['ns_status'] == 'FAIL':
            ns_status_cell.fill = RED_FILL

        # MX
        mx_status_cell = ws.cell(row=row_idx, column=6, value=result['mx_status'])
        ws.cell(row=row_idx, column=7, value=result['mx_note'])
        ws.cell(row=row_idx, column=8, value=result['mx_detail'])
        if result['mx_status'] == 'PASS':
            mx_status_cell.fill = GREEN_FILL
        elif result['mx_status'] == 'FAIL':
            mx_status_cell.fill = RED_FILL
        elif result['mx_status'] == 'WARN':
            mx_status_cell.fill = YELLOW_FILL

        # SPF
        spf_status_cell = ws.cell(row=row_idx, column=9, value=result['spf_status'])
        ws.cell(row=row_idx, column=10, value=result['spf_note'])
        ws.cell(row=row_idx, column=11, value=result['spf_detail'])
        if result['spf_status'] == 'PASS':
            spf_status_cell.fill = GREEN_FILL
        elif result['spf_status'] == 'FAIL':
            spf_status_cell.fill = RED_FILL
        elif result['spf_status'] == 'WARN':
            spf_status_cell.fill = YELLOW_FILL

        # DKIM
        dkim_status_cell = ws.cell(row=row_idx, column=12, value=result['dkim_status'])
        ws.cell(row=row_idx, column=13, value=result['dkim_note'])
        ws.cell(row=row_idx, column=14, value=result['dkim_detail'])
        if result['dkim_status'] == 'PASS':
            dkim_status_cell.fill = GREEN_FILL
        elif result['dkim_status'] == 'FAIL':
            dkim_status_cell.fill = RED_FILL
        elif result['dkim_status'] == 'WARN':
            dkim_status_cell.fill = YELLOW_FILL

        # DMARC
        dmarc_status_cell = ws.cell(row=row_idx, column=15, value=result['dmarc_status'])
        ws.cell(row=row_idx, column=16, value=result['dmarc_note'])
        ws.cell(row=row_idx, column=17, value=result['dmarc_detail'])
        if result['dmarc_status'] == 'PASS':
            dmarc_status_cell.fill = GREEN_FILL
        elif result['dmarc_status'] == 'FAIL':
            dmarc_status_cell.fill = RED_FILL
        elif result['dmarc_status'] == 'WARN':
            dmarc_status_cell.fill = YELLOW_FILL

        # Audit date
        ws.cell(row=row_idx, column=18, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

    # Adjust column widths
    ws.column_dimensions['A'].width = 40  # Domain
    ws.column_dimensions['B'].width = 15  # Overall Status
    ws.column_dimensions['C'].width = 12  # NS Status
    ws.column_dimensions['D'].width = 30  # NS Note
    ws.column_dimensions['E'].width = 50  # NS Details
    ws.column_dimensions['F'].width = 12  # MX Status
    ws.column_dimensions['G'].width = 30  # MX Note
    ws.column_dimensions['H'].width = 50  # MX Details
    ws.column_dimensions['I'].width = 12  # SPF Status
    ws.column_dimensions['J'].width = 30  # SPF Note
    ws.column_dimensions['K'].width = 80  # SPF Details (full record)
    ws.column_dimensions['L'].width = 12  # DKIM Status
    ws.column_dimensions['M'].width = 30  # DKIM Note
    ws.column_dimensions['N'].width = 50  # DKIM Details
    ws.column_dimensions['O'].width = 12  # DMARC Status
    ws.column_dimensions['P'].width = 30  # DMARC Note
    ws.column_dimensions['Q'].width = 80  # DMARC Details (full record)
    ws.column_dimensions['R'].width = 18  # Audit Date

    # Save workbook
    wb.save(output_file)
    print(f"\n✅ Results saved to: {output_file}")

def main():
    input_file = '/Users/YOUR_USERNAME/Library/CloudStorage/OneDrive-YOUR_ORG/Documents/Claude/Route53 domains - 2025.10.10 1.xlsx'
    output_file = '/Users/YOUR_USERNAME/Library/CloudStorage/OneDrive-YOUR_ORG/Documents/Claude/Route53 domains - DNS Audit Results.xlsx'

    # Load domains
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    domains = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            domain = str(row[0]).strip().rstrip('.')
            domains.append(domain)

    print(f"📋 Found {len(domains)} domains to audit\n")

    # Run audit
    auditor = DNSAuditor()
    results = []

    for idx, domain in enumerate(domains, start=1):
        print(f"[{idx}/{len(domains)}] ", end="")
        result = auditor.audit_domain(domain)
        results.append(result)

    # Update Excel
    print(f"\n📝 Updating Excel with results...")
    update_excel_with_results(input_file, output_file, results)

    # Summary
    overall_pass = sum(1 for r in results if r['overall_status'] == 'PASS')
    overall_warn = sum(1 for r in results if r['overall_status'] == 'WARN')
    overall_fail = sum(1 for r in results if r['overall_status'] == 'FAIL')
    overall_error = sum(1 for r in results if r['overall_status'] == 'ERROR')

    print(f"\n📊 DNS Audit Summary:")
    print(f"   ✅ PASS:  {overall_pass}/{len(domains)} ({overall_pass*100//len(domains)}%)")
    print(f"   ⚠️  WARN:  {overall_warn}/{len(domains)} ({overall_warn*100//len(domains)}%)")
    print(f"   ❌ FAIL:  {overall_fail}/{len(domains)} ({overall_fail*100//len(domains)}%)")
    print(f"   🔴 ERROR: {overall_error}/{len(domains)} ({overall_error*100//len(domains)}%)")

    # Breakdown
    spf_fail = sum(1 for r in results if r['spf_status'] == 'FAIL')
    dkim_fail = sum(1 for r in results if r['dkim_status'] == 'FAIL')
    dmarc_fail = sum(1 for r in results if r['dmarc_status'] == 'FAIL')
    mx_fail = sum(1 for r in results if r['mx_status'] == 'FAIL')

    print(f"\n🔍 Issues Breakdown:")
    print(f"   SPF missing/failed:   {spf_fail} domains")
    print(f"   DKIM missing/failed:  {dkim_fail} domains")
    print(f"   DMARC missing/failed: {dmarc_fail} domains")
    print(f"   MX missing/failed:    {mx_fail} domains")

if __name__ == '__main__':
    main()
