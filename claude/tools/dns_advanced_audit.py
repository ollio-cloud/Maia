#!/usr/bin/env python3
"""
Advanced DNS Security & Deliverability Audit Tool
Integrates multiple open-source tools for comprehensive DNS analysis
"""

import dns.resolver
import dns.exception
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import subprocess
import json
import re
import sys
from typing import Dict, List, Tuple, Optional
import requests

# Color fills for Excel
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
BOLD_FONT = Font(bold=True)

class AdvancedDNSAuditor:
    def __init__(self):
        self.resolver = dns.resolver.Resolver()
        self.resolver.timeout = 5
        self.resolver.lifetime = 5

    def check_dnssec(self, domain: str) -> Tuple[str, str, str]:
        """Check DNSSEC configuration"""
        try:
            # Try to get DNSKEY records
            answers = self.resolver.resolve(domain, 'DNSKEY')
            if answers:
                return "PASS", f"DNSSEC enabled ({len(answers)} keys)", f"{len(answers)} DNSKEY records found"
        except dns.resolver.NoAnswer:
            return "WARN", "DNSSEC not configured", "No DNSKEY records found"
        except dns.resolver.NXDOMAIN:
            return "FAIL", "Domain does not exist", ""
        except Exception as e:
            return "WARN", "DNSSEC check inconclusive", str(e)[:50]

    def check_caa(self, domain: str) -> Tuple[str, str, str]:
        """Check CAA records for certificate authority authorization"""
        try:
            answers = self.resolver.resolve(domain, 'CAA')
            caa_records = []
            for rdata in answers:
                caa_records.append(f"{rdata.flags} {rdata.tag} {rdata.value}")

            if caa_records:
                return "PASS", f"CAA configured ({len(caa_records)} records)", "; ".join(caa_records)[:200]
            else:
                return "WARN", "No CAA records", ""
        except dns.resolver.NoAnswer:
            return "WARN", "No CAA records (any CA can issue)", "Recommended: Add CAA records"
        except dns.resolver.NXDOMAIN:
            return "FAIL", "Domain does not exist", ""
        except Exception as e:
            return "INFO", "CAA check inconclusive", str(e)[:50]

    def check_mta_sts(self, domain: str) -> Tuple[str, str, str]:
        """Check MTA-STS policy for SMTP TLS enforcement"""
        mta_sts_domain = f"_mta-sts.{domain}"
        try:
            answers = self.resolver.resolve(mta_sts_domain, 'TXT')
            for rdata in answers:
                record = str(rdata).strip('"')
                if 'v=STSv1' in record:
                    # Try to fetch the policy file
                    try:
                        policy_url = f"https://mta-sts.{domain}/.well-known/mta-sts.txt"
                        response = requests.get(policy_url, timeout=5)
                        if response.status_code == 200:
                            return "PASS", "MTA-STS fully configured", record[:100]
                        else:
                            return "WARN", "MTA-STS DNS record exists but policy file missing", record[:100]
                    except:
                        return "WARN", "MTA-STS DNS record exists but policy unreachable", record[:100]

            return "WARN", "No MTA-STS record", ""
        except (dns.resolver.NoAnswer, dns.resolver.NXDOMAIN):
            return "WARN", "No MTA-STS (SMTP TLS not enforced)", "Recommended for email security"
        except Exception as e:
            return "INFO", "MTA-STS check inconclusive", str(e)[:50]

    def check_tls_rpt(self, domain: str) -> Tuple[str, str, str]:
        """Check TLS-RPT record for TLS reporting"""
        tls_rpt_domain = f"_smtp._tls.{domain}"
        try:
            answers = self.resolver.resolve(tls_rpt_domain, 'TXT')
            for rdata in answers:
                record = str(rdata).strip('"')
                if 'v=TLSRPTv1' in record:
                    return "PASS", "TLS-RPT configured", record[:100]

            return "WARN", "No TLS-RPT record", ""
        except (dns.resolver.NoAnswer, dns.resolver.NXDOMAIN):
            return "WARN", "No TLS-RPT (TLS reporting disabled)", "Recommended with MTA-STS"
        except Exception as e:
            return "INFO", "TLS-RPT check inconclusive", str(e)[:50]

    def check_bimi(self, domain: str) -> Tuple[str, str, str]:
        """Check BIMI record for brand logo in email"""
        bimi_domain = f"default._bimi.{domain}"
        try:
            answers = self.resolver.resolve(bimi_domain, 'TXT')
            for rdata in answers:
                record = str(rdata).strip('"')
                if 'v=BIMI1' in record:
                    return "PASS", "BIMI configured (brand logo)", record[:100]

            return "INFO", "No BIMI record", ""
        except (dns.resolver.NoAnswer, dns.resolver.NXDOMAIN):
            return "INFO", "No BIMI (optional brand enhancement)", ""
        except Exception as e:
            return "INFO", "BIMI check inconclusive", str(e)[:50]

    def check_subdomain_takeover(self, domain: str) -> Tuple[str, str, str]:
        """Check for potential subdomain takeover vulnerabilities"""
        vulnerable_patterns = [
            'github.io',
            'herokuapp.com',
            'azurewebsites.net',
            'cloudapp.net',
            's3.amazonaws.com',
            'elasticbeanstalk.com',
            'cloudfront.net',
            'wordpress.com',
            'pantheonsite.io',
            'zendesk.com',
            'desk.com',
            'statuspage.io',
            'uservoice.com',
            'surge.sh',
            'bitbucket.io',
            'freshdesk.com',
            'ghost.io',
            'helpscoutdocs.com',
            'helpjuice.com',
            'readme.io',
            'statuspage.io',
            'tumblr.com',
            'unbounce.com',
            'campaignmonitor.com',
            'acquia-sites.com',
        ]

        try:
            # Check common subdomains
            common_subdomains = ['www', 'blog', 'shop', 'api', 'dev', 'staging', 'test', 'mail', 'ftp']
            vulnerable_count = 0
            vulnerable_subdomains = []

            for subdomain in common_subdomains:
                try:
                    full_domain = f"{subdomain}.{domain}"
                    answers = self.resolver.resolve(full_domain, 'CNAME')
                    for rdata in answers:
                        cname_target = str(rdata.target).lower()
                        for pattern in vulnerable_patterns:
                            if pattern in cname_target:
                                vulnerable_count += 1
                                vulnerable_subdomains.append(f"{subdomain}→{cname_target[:30]}")
                                break
                except:
                    continue

            if vulnerable_count > 0:
                return "FAIL", f"Potential takeover risk ({vulnerable_count} subdomains)", ", ".join(vulnerable_subdomains[:3])
            else:
                return "PASS", "No obvious takeover risks detected", f"Checked {len(common_subdomains)} common subdomains"

        except Exception as e:
            return "INFO", "Subdomain takeover check inconclusive", str(e)[:50]

    def check_ptr_records(self, domain: str) -> Tuple[str, str, str]:
        """Check PTR (reverse DNS) records for mail servers"""
        try:
            # Get MX records first
            mx_answers = self.resolver.resolve(domain, 'MX')
            mx_hosts = [str(rdata.exchange).rstrip('.') for rdata in mx_answers]

            ptr_results = []
            for mx_host in mx_hosts[:3]:  # Check first 3 MX records
                try:
                    # Get A record for MX host
                    a_answers = self.resolver.resolve(mx_host, 'A')
                    for a_rdata in a_answers:
                        ip = str(a_rdata)
                        # Try reverse DNS lookup
                        try:
                            ptr_answer = self.resolver.resolve_address(ip)
                            ptr_results.append(f"{mx_host[:20]}→{ip}→PTR:OK")
                        except:
                            ptr_results.append(f"{mx_host[:20]}→{ip}→PTR:MISSING")
                except:
                    continue

            if not ptr_results:
                return "INFO", "No MX records to check PTR", ""

            missing_count = sum(1 for r in ptr_results if 'MISSING' in r)
            if missing_count > 0:
                return "WARN", f"PTR records missing ({missing_count}/{len(ptr_results)})", ", ".join(ptr_results)[:200]
            else:
                return "PASS", f"All PTR records configured ({len(ptr_results)})", ", ".join(ptr_results)[:200]

        except dns.resolver.NoAnswer:
            return "INFO", "No MX records to check PTR", ""
        except Exception as e:
            return "INFO", "PTR check inconclusive", str(e)[:50]

    def check_dns_propagation(self, domain: str) -> Tuple[str, str, str]:
        """Check DNS propagation across multiple public resolvers"""
        public_resolvers = [
            '8.8.8.8',      # Google
            '1.1.1.1',      # Cloudflare
            '208.67.222.222', # OpenDNS
        ]

        ns_results = {}
        for resolver_ip in public_resolvers:
            temp_resolver = dns.resolver.Resolver()
            temp_resolver.nameservers = [resolver_ip]
            temp_resolver.timeout = 3
            temp_resolver.lifetime = 3

            try:
                answers = temp_resolver.resolve(domain, 'NS')
                ns_set = frozenset(str(rdata) for rdata in answers)
                ns_results[resolver_ip] = ns_set
            except:
                ns_results[resolver_ip] = None

        # Check consistency
        valid_results = [ns_set for ns_set in ns_results.values() if ns_set is not None]

        if not valid_results:
            return "FAIL", "DNS not resolvable from public resolvers", ""

        if len(set(valid_results)) == 1:
            return "PASS", f"DNS consistent across {len(valid_results)} resolvers", ""
        else:
            return "WARN", "DNS inconsistent across resolvers", f"Different NS records from different resolvers"

    def check_ttl_values(self, domain: str) -> Tuple[str, str, str]:
        """Check TTL values for DNS records"""
        try:
            # Check A record TTL
            answers = self.resolver.resolve(domain, 'A')
            ttl = answers.rrset.ttl

            if ttl < 300:
                return "WARN", f"Very low TTL ({ttl}s)", "May cause excessive DNS queries"
            elif ttl > 86400:
                return "WARN", f"Very high TTL ({ttl}s)", "Slow propagation for changes"
            else:
                return "PASS", f"Appropriate TTL ({ttl}s)", "Good balance"

        except dns.resolver.NoAnswer:
            return "INFO", "No A record to check TTL", ""
        except Exception as e:
            return "INFO", "TTL check inconclusive", str(e)[:50]

def audit_domain_advanced(domain: str) -> Dict:
    """Run advanced DNS audit for domain"""
    domain = domain.strip().rstrip('.')
    print(f"Advanced auditing: {domain}")

    auditor = AdvancedDNSAuditor()

    # Run all advanced checks
    dnssec_status, dnssec_note, dnssec_detail = auditor.check_dnssec(domain)
    caa_status, caa_note, caa_detail = auditor.check_caa(domain)
    mta_sts_status, mta_sts_note, mta_sts_detail = auditor.check_mta_sts(domain)
    tls_rpt_status, tls_rpt_note, tls_rpt_detail = auditor.check_tls_rpt(domain)
    bimi_status, bimi_note, bimi_detail = auditor.check_bimi(domain)
    subdomain_status, subdomain_note, subdomain_detail = auditor.check_subdomain_takeover(domain)
    ptr_status, ptr_note, ptr_detail = auditor.check_ptr_records(domain)
    propagation_status, propagation_note, propagation_detail = auditor.check_dns_propagation(domain)
    ttl_status, ttl_note, ttl_detail = auditor.check_ttl_values(domain)

    # Calculate security score
    security_checks = [dnssec_status, caa_status, mta_sts_status, tls_rpt_status, subdomain_status]
    pass_count = sum(1 for s in security_checks if s == 'PASS')
    warn_count = sum(1 for s in security_checks if s == 'WARN')
    fail_count = sum(1 for s in security_checks if s == 'FAIL')

    if fail_count > 0:
        security_score = "HIGH RISK"
    elif warn_count > 2:
        security_score = "MEDIUM RISK"
    elif pass_count >= 3:
        security_score = "LOW RISK"
    else:
        security_score = "MEDIUM RISK"

    return {
        'domain': domain,
        'security_score': security_score,
        'dnssec_status': dnssec_status,
        'dnssec_note': dnssec_note,
        'dnssec_detail': dnssec_detail,
        'caa_status': caa_status,
        'caa_note': caa_note,
        'caa_detail': caa_detail,
        'mta_sts_status': mta_sts_status,
        'mta_sts_note': mta_sts_note,
        'mta_sts_detail': mta_sts_detail,
        'tls_rpt_status': tls_rpt_status,
        'tls_rpt_note': tls_rpt_note,
        'tls_rpt_detail': tls_rpt_detail,
        'bimi_status': bimi_status,
        'bimi_note': bimi_note,
        'bimi_detail': bimi_detail,
        'subdomain_status': subdomain_status,
        'subdomain_note': subdomain_note,
        'subdomain_detail': subdomain_detail,
        'ptr_status': ptr_status,
        'ptr_note': ptr_note,
        'ptr_detail': ptr_detail,
        'propagation_status': propagation_status,
        'propagation_note': propagation_note,
        'propagation_detail': propagation_detail,
        'ttl_status': ttl_status,
        'ttl_note': ttl_note,
        'ttl_detail': ttl_detail,
    }

def main():
    # Test with a few domains first
    test_domains = ['google.com', 'github.com', 'microsoft.com']

    print("🔬 Advanced DNS Security Audit\n")

    for domain in test_domains:
        result = audit_domain_advanced(domain)
        print(f"\n{domain}:")
        print(f"  Security Score: {result['security_score']}")
        print(f"  DNSSEC: {result['dnssec_status']} - {result['dnssec_note']}")
        print(f"  CAA: {result['caa_status']} - {result['caa_note']}")
        print(f"  MTA-STS: {result['mta_sts_status']} - {result['mta_sts_note']}")
        print(f"  Subdomain Takeover: {result['subdomain_status']} - {result['subdomain_note']}")

if __name__ == '__main__':
    main()
